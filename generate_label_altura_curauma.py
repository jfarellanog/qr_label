import qrcode
from PIL import Image, ImageDraw, ImageFont
import numpy as np
import pandas as pd
import xlsxwriter
import openpyxl as xl
from datetime import datetime
from openpyxl import load_workbook

excel_file = r'C:\Users\jarellan\Desktop\qr_label\IN\Ubicaciones'
ubicaciones = pd.read_excel(excel_file+".xlsx")

# My image is a 200x374 jpeg that is 102kb large
logo_ccu = Image.open('logo_ccu.png')

# I downsize the image with an ANTIALIAS filter (gives the highest quality)
logo_ccu = logo_ccu.resize((150, 140), Image.ANTIALIAS)
logo_ccu.save("logo_ccu_escalado.png", quality=95)
# The saved downsized image size is 24.8kb
logo_ccu.save("logo_ccu_escalado_opt.png", optimize=True, quality=95)
# The saved downsized image size is 22.9kb
ubicacion = []
fecha = datetime.today().strftime("%d-%m-%Y")
k=0
df = pd.read_excel('C:\\Users\\jarellan\\Desktop\\qr_label\\OUT\\CURAUMA\\Etiquetas_Impresas_CURAUMA.xlsx')
df = df.dropna()
print(df)
if len(df)>0:
    last_row = int(len(df))
    id_etiqueta = int(df['Id'].max())
else:
    last_row = 0
    id_etiqueta = 0
for index, row in ubicaciones.iloc[0:].iterrows():
        images = []
        # Con crop corto la imagen
        # face = Image.open('logo_ccu.png')
        code = row['Posicion']
        code2 = '-01-01'
        img = Image.new('RGB', (950,840), color=(255, 255, 255))
        img2 = Image.new('RGB', (650, 400), color=(255, 255, 255))
        img3 = Image.new('RGB', (650, 400), color=(255, 255, 255))
        fnt = ImageFont.truetype(r'C:\Users\jarellan\Desktop\qr_label\IN\Futura-CondensedLight.ttf', 300)
        fnt2 = ImageFont.truetype(r'G:\Unidades compartidas\Centro Excelencia Logística CCU S.A\QR\ariali.ttf', 200)
        fnt3 = ImageFont.truetype(r'C:\Users\jarellan\Desktop\qr_label\IN\Futura-CondensedLight.ttf', 350)
        fnt4 = ImageFont.truetype(r'C:\Users\jarellan\Desktop\qr_label\IN\Futura-CondensedLight.ttf', 300)
        d = ImageDraw.Draw(img)
        d2 = ImageDraw.Draw(img2)
        d3 = ImageDraw.Draw(img3)
        title = 'PICKING'
        #d.line((70, 45, 1000, 45), fill=(0, 0, 0), width=1)
        #d.text((160, 50), title, font=fnt2, fill=(0, 0, 0))
        d.line((70, 60, 950, 60), fill=(0, 0, 0), width=1)
        d2.text((10, 10), code, font=fnt, fill=(0, 0, 0))
        d3.text((10, 10), code2, font=fnt, fill=(0, 0, 0))
        d.line((70, 800, 950, 800), fill=(0, 0, 0), width=1)
        # d.line((70, 190, 750, 190), fill=(0, 0, 0), width=1)
        r = int(row[1])
        c = int(row[2])
        t = int(row[3])

        if r < 10:
            rectangle = '0' + str(r)
        else:
            rectangle = str(r)
        if c < 10:
            circle = '0' + str(c)
        else:
            circle = str(c)
        if t < 10:
            triangle = '0' + str(t)
        else:
            triangle = str(t)

        d.rectangle((80, 90, 400, 420), fill="black", outline="black")
        d.text((90, 100), rectangle, font=fnt3, fill=(255, 255, 255))
        d.ellipse((600, 70, 950, 450), fill="black", outline="black")
        d.text((630, 100), circle, font=fnt3, fill=(255, 255, 255))
        d.polygon([(510, 350), (300, 770), (760, 770)], fill="black", outline="black")
        d.text((400,520), triangle, font=fnt4, fill=(255, 255, 255))

        qr_big = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=30,border=0)
        qr_big.add_data(code)
        qr_big.make()
        img_qr_big = qr_big.make_image().convert('RGB')
        pos = ((img_qr_big.size[0] - logo_ccu.size[0]) // 2, (img_qr_big.size[1] - logo_ccu.size[1]) // 2)
        img_qr_big.paste(logo_ccu, pos)

        imgs = [img_qr_big]
        min_shape = sorted([(np.sum(i.size), i.size) for i in imgs])[0][1]
        imgs_comb = np.hstack((np.asarray(i.resize(min_shape)) for i in imgs))

        # save that beautiful picture
        imgs_comb = Image.fromarray(imgs_comb)
        id_etiqueta += 1
        imgs_comb.save('C:\\Users\\jarellan\\Desktop\\qr_label\\OUT\\CURAUMA\\qr' + row['Posicion'] + '_' + str(
            id_etiqueta + k) + '.jpg')
        print('Impresión de qr' + row['Posicion'] + '_' + str(k))

        imgs = [img]
        min_shape = sorted([(np.sum(i.size), i.size) for i in imgs])[0][1]
        imgs_comb = np.hstack((np.asarray(i.resize(min_shape)) for i in imgs))

        # save that beautiful picture
        imgs_comb = Image.fromarray(imgs_comb)
        id_etiqueta += 1
        imgs_comb.save('C:\\Users\\jarellan\\Desktop\\qr_label\\OUT\\CURAUMA\\digitos' + row['Posicion'] + '_' + str(
            id_etiqueta + k) + '.jpg')
        print('Impresión de digitos' + row['Posicion'] + '_' + str(k))

        imgs = [img2]
        min_shape = sorted([(np.sum(i.size), i.size) for i in imgs])[0][1]
        imgs_comb = np.hstack((np.asarray(i.resize(min_shape)) for i in imgs))

        # save that beautiful picture
        id_etiqueta += 1
        imgs_comb.save('C:\\Users\\jarellan\\Desktop\\qr_label\\OUT\\CURAUMA\\codigo' + row['Posicion'] + '_' + str(
            id_etiqueta + k) + '.jpg')
        print('Impresión de codigo parte 1' + row['Posicion'] + '_' + str(k))

        imgs = [img3]
        min_shape = sorted([(np.sum(i.size), i.size) for i in imgs])[0][1]
        imgs_comb = np.hstack((np.asarray(i.resize(min_shape)) for i in imgs))

        # save that beautiful picture
        imgs_comb = Image.fromarray(imgs_comb)
        id_etiqueta += 1
        imgs_comb = Image.fromarray(imgs_comb)
        imgs_comb.save('C:\\Users\\jarellan\\Desktop\\qr_label\\OUT\\CURAUMA\\codigo' + row['Posicion'] + '_' + str(
                id_etiqueta + k) + '.jpg')
        print('Impresión de codigo parte 2' + row['Posicion'] + '_' + str(k))

        df.loc[last_row]=[str(code),str(rectangle),str(circle),str(triangle),int(id_etiqueta + k),row['Posicion'] + '_' + str(id_etiqueta + k)]
        print(df)
        k += 1




path = 'C:\\Users\\jarellan\\Desktop\\qr_label\\OUT\\CURAUMA\\Etiquetas_Impresas_CURAUMA.xlsx'
#book = load_workbook(filename=path)
#writer = pd.ExcelWriter(path, engine = 'openpyxl')
#writer.book = book
#df.to_excel(writer,sheet_name='Impresiones',index=False)
#writer.save()
#writer.close()

book = load_workbook(path)
writer = pd.ExcelWriter(path, engine='openpyxl')
writer.book = book
df.to_excel(writer, sheet_name='Impresiones',index = False)
writer.save()