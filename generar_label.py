import qrcode
from PIL import Image, ImageDraw, ImageFont
import numpy as np
import pandas as pd
import xlsxwriter
import openpyxl as xl
from datetime import datetime

excel_file = r'C:\Users\jarellan\Desktop\qr_label\IN\Ubicaciones'
ubicaciones = pd.read_excel(excel_file+".xlsx")

# My image is a 200x374 jpeg that is 102kb large
logo_ccu = Image.open('logo_ccu.png')

# I downsize the image with an ANTIALIAS filter (gives the highest quality)
logo_ccu = logo_ccu.resize((80, 70), Image.ANTIALIAS)
logo_ccu.save("logo_ccu_escalado.png", quality=95)
# The saved downsized image size is 24.8kb
logo_ccu.save("logo_ccu_escalado_opt.png", optimize=True, quality=95)
# The saved downsized image size is 22.9kb
ubicacion = []
fecha = datetime.today().strftime("%d-%m-%Y")
c=0
wb = xl.load_workbook(filename='C:\\Users\\jarellan\\Desktop\\qr_label\\OUT\\Etiquetas_Impresas.xlsx')
df = pd.read_excel('C:\\Users\\jarellan\\Desktop\\qr_label\\OUT\\Etiquetas_Impresas.xlsx', index_col=0)
ws = wb.worksheets[0]
print(df)
if len(df)>0:
    last_row = int(ws.max_row)
    id_etiqueta = int(df['Id'].max())
else:
    last_row = 0
    id_etiqueta = 0
v=1
if v==1:
    for index, row in ubicaciones.iloc[0:].iterrows():
        images = []
# Con crop corto la imagen
#face = Image.open('logo_ccu.png')
        code = row['Posicion']
        img = Image.new('RGB', (755, 300), color=(255, 255, 255))
        fnt = ImageFont.truetype(r'G:\Unidades compartidas\Centro Excelencia Logística CCU S.A\QR\ariali.ttf', 100)
        fnt2 = ImageFont.truetype(r'G:\Unidades compartidas\Centro Excelencia Logística CCU S.A\QR\arialbi.ttf', 50)
        fnt3 = ImageFont.truetype(r'G:\Unidades compartidas\Centro Excelencia Logística CCU S.A\QR\arialbi.ttf', 50)
        d = ImageDraw.Draw(img)
        title = 'PICKING VALORADO'
        d.line((70,100,750,100),fill=(0, 0, 0),width=1)
        d.line((70, 45, 750, 45), fill=(0, 0, 0), width=1)
        d.text((140,50), title, font=fnt2, fill=(0, 0, 0))
        d.text((70,90), code, font=fnt, fill=(0, 0, 0))
        d.line((70,270,750,270),fill=(0, 0, 0),width=1)
        #d.line((70, 190, 750, 190), fill=(0, 0, 0), width=1)
        r = int(row[1])
        c = int(row[2])
        t = int(row[3])

        if r<10:
            rectangle = '0'+str(r)
        else:
            rectangle=str(r)
        if c<10:
            circle = '0'+str(c)
        else:
            circle=str(c)
        if t<10:
            triangle = '0'+str(t)
        else:
            triangle=str(t)

        d.rectangle((70, 190, 220, 260), fill="white",outline="black")
        d.text((100,200), rectangle, font=fnt3,fill=(0, 0, 0))
        d.ellipse((300, 190, 470, 260), fill="white",outline="black")
        d.text((350,200), circle, font=fnt3,fill=(0, 0, 0))
        d.polygon([(650, 190),(580, 260),(730, 260)], fill="white",outline="black")
        d.text((620,210), triangle, font=fnt3,fill=(0, 0, 0))


        qr_big = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H,box_size=15)
        qr_big.add_data(code)
        qr_big.make()
        img_qr_big = qr_big.make_image().convert('RGB')
        pos = ((img_qr_big.size[0] - logo_ccu.size[0]) // 2, (img_qr_big.size[1] - logo_ccu.size[1]) // 2)
        img_qr_big.paste(logo_ccu, pos)

        imgs = [img,img_qr_big]
        min_shape = sorted([(np.sum(i.size), i.size) for i in imgs])[0][1]
        imgs_comb = np.hstack((np.asarray(i.resize(min_shape)) for i in imgs))

        # save that beautiful picture
        imgs_comb = Image.fromarray(imgs_comb)
        imgs_comb.save( 'C:\\Users\\jarellan\\Desktop\\qr_label\\OUT\\etiqueta'+row['Posicion']+'_'+str(id_etiqueta+c)+'.jpg' )
        print('Impresión de '+row['Posicion']+'_'+str(c))
        ws.cell(row=c + last_row + 1, column=1).value = str(code)
        ws.cell(row=c + last_row + 1, column=2).value = str(circle)
        ws.cell(row=c + last_row + 1, column=3).value = rectangle
        ws.cell(row=c + last_row + 1, column=4).value = triangle
        ws.cell(row=c + last_row + 1, column=5).value = int(id_etiqueta+c)
        ws.cell(row=c + last_row + 1, column=6).value = row['Posicion']+'_'+str(id_etiqueta+c)
        c+=1
else:
    for index, row in ubicaciones.iloc[0:].iterrows():
        images = []
# Con crop corto la imagen
#face = Image.open('logo_ccu.png')
        code = row['Posicion']
        img = Image.new('RGB', (755, 300), color=(255, 255, 255))
        fnt = ImageFont.truetype(r'G:\Unidades compartidas\Centro Excelencia Logística CCU S.A\QR\arial.ttf', 100)
        fnt2 = ImageFont.truetype(r'G:\Unidades compartidas\Centro Excelencia Logística CCU S.A\QR\arial.ttf', 50)
        fnt3 = ImageFont.truetype(r'G:\Unidades compartidas\Centro Excelencia Logística CCU S.A\QR\arial.ttf', 50)
        d = ImageDraw.Draw(img)
        title = 'PICKING'
        d.text((270,50), title, font=fnt2, fill=(0, 0, 0))
        d.text((70,90), code, font=fnt, fill=(0, 0, 0))
        rectangle = str(int(row[1]))
        circle = str(int(row[2]))
        triangle = str(int(row[3]))

        d.rectangle((70, 190, 220, 260), fill="white",outline="black")
        d.text((100,200), rectangle, font=fnt3,fill=(0, 0, 0))
        d.ellipse((300, 190, 470, 260), fill="white",outline="black")
        d.text((350,200), circle, font=fnt3,fill=(0, 0, 0))
        d.polygon([(650, 190),(580, 260),(730, 260)], fill="white",outline="black")
        d.text((620,210), triangle, font=fnt3,fill=(0, 0, 0))


        qr_big = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H,box_size=15)
        qr_big.add_data(code)
        qr_big.make()
        img_qr_big = qr_big.make_image().convert('RGB')
        pos = ((img_qr_big.size[0] - logo_ccu.size[0]) // 2, (img_qr_big.size[1] - logo_ccu.size[1]) // 2)
        img_qr_big.paste(logo_ccu, pos)

        imgs = [img,img_qr_big]
        min_shape = sorted([(np.sum(i.size), i.size) for i in imgs])[0][1]
        imgs_comb = np.hstack((np.asarray(i.resize(min_shape)) for i in imgs))

        # save that beautiful picture
        imgs_comb = Image.fromarray(imgs_comb)
        imgs_comb.save( 'C:\\Users\\jarellan\\Desktop\\qr_label\\OUT\\etiqueta'+row['Posicion']+'_'+str(id_etiqueta+c)+'.jpg' )
        print('Impresión de '+row['Posicion']+'_'+str(c))
        ws.cell(row=c + last_row + 1, column=1).value = str(code)
        ws.cell(row=c + last_row + 1, column=2).value = str(circle)
        ws.cell(row=c + last_row + 1, column=3).value = rectangle
        ws.cell(row=c + last_row + 1, column=4).value = triangle
        ws.cell(row=c + last_row + 1, column=5).value = int(id_etiqueta+c)
        ws.cell(row=c + last_row + 1, column=6).value = row['Posicion']+'_'+str(id_etiqueta+c)
        c+=1
wb.save(str('C:\\Users\\jarellan\\Desktop\\qr_label\\OUT\\Etiquetas_Impresas.xlsx'))
